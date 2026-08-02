import re
from pathlib import Path
import openpyxl

from app.database import Base, SessionLocal, engine
from app.matching import EMBEDDING_MODEL, embed_text
from app.models import LciProcess

Base.metadata.create_all(bind=engine)

EXCEL_PATH = Path(__file__).resolve().parent.parent / "GHG_Calculator_RECTIFIED_v6.xlsx"


def run_seed():
    if not EXCEL_PATH.exists():
        print(f"Error: Could not find Excel file at {EXCEL_PATH}")
        return

    db = SessionLocal()
    try:
        existing_count = db.query(LciProcess).filter(LciProcess.database_source == "India_GHG_Factors").count()
        if existing_count > 0:
            print(f"Database already contains {existing_count} India_GHG_Factors rows; skipping seed.")
            return

        wb_val = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        wb_form = openpyxl.load_workbook(EXCEL_PATH, data_only=False)

        db_sheet = wb_val['DB_Master']
        db_rows = list(db_sheet.iter_rows(values_only=True))

        db_data = []
        for row in db_rows[1:]:
            if row[0] is not None:
                key = str(row[0]).strip()
                factor = float(row[1]) if row[1] is not None else 0.0
                source_notes = str(row[2]).strip() if row[2] is not None else ""
                db_data.append({
                    'key': key,
                    'factor': factor,
                    'source_notes': source_notes
                })

        ghg_form_sheet = wb_form['GHG_Master_Calculator']
        ghg_val_sheet = wb_val['GHG_Master_Calculator']

        ghg_form_rows = list(ghg_form_sheet.iter_rows(values_only=True))
        ghg_val_rows = list(ghg_val_sheet.iter_rows(values_only=True))

        key_to_process = {}

        for idx, f_row in enumerate(ghg_form_rows):
            v_row = ghg_val_rows[idx]
            for cell_val in f_row:
                if cell_val and isinstance(cell_val, str) and "=VLOOKUP(" in cell_val:
                    m = re.search(r'=VLOOKUP\("([^"]+)"', cell_val)
                    if m:
                        k = m.group(1)
                        process_name = str(v_row[0]).strip() if v_row[0] is not None else ""
                        unit = str(v_row[2]).strip() if v_row[2] is not None else "unit"
                        if k not in key_to_process:
                            key_to_process[k] = {'name': process_name, 'unit': unit}

        added_count = 0
        for item in db_data:
            key = item['key']
            proc_info = key_to_process.get(key, {'name': key, 'unit': 'unit'})
            process_name = proc_info['name']
            reference_unit = proc_info['unit']
            notes = item['source_notes']

            if "PLACEHOLDER" in notes:
                status = "placeholder"
            elif "uplift" in notes.lower() or "EPA GHG Equivalencies" in notes:
                status = "uplifted"
            elif "Cross-validated" in notes or "RECONCILED" in notes:
                status = "proxy"
            else:
                status = "clean"

            embed_input = f"{process_name} {notes}"
            vector = embed_text(embed_input)

            process = LciProcess(
                process_uuid=key,
                database_source="India_GHG_Factors",
                database_version="v6",
                process_name=process_name,
                reference_product=process_name,
                reference_unit=reference_unit,
                geography="IN",
                system_model="Direct Factor",
                description=notes,
                emission_factor=item['factor'],
                emission_factor_source=notes,
                data_quality_status=status,
                embedding=vector,
                embedding_model=EMBEDDING_MODEL,
                is_active=True
            )
            db.add(process)
            added_count += 1

        db.commit()
        print(f"Successfully seeded {added_count} India_GHG_Factors processes.")

    finally:
        db.close()


if __name__ == "__main__":
    run_seed()
